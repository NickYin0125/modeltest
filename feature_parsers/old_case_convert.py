import textwrap
from dataclasses import dataclass
from pathlib import Path
from typing import List, Tuple, Dict, Set, Optional

from pytest_bdd.gherkin_parser import Background
from pytest_bdd.parser import FeatureParser, ScenarioTemplate, Step
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


# -----------------------------------------------------------------------------
# Dataclass representing one Excel row
# -----------------------------------------------------------------------------

@dataclass
class ParsedTestCase:
    case_name: str
    test_data: Dict[str, str]
    ordered_test_steps: List[Tuple[int, str]]  # Given + When with continuous seq
    expected_results: Dict[int, str]           # key = When‑seq
    fr_id: str                                # formatted ODPCL tags (one per line)
    ccd_batch_status: str                      # value from CCDBatchStatus tag
    ncd_batch_status: str                      # value from NCDBatchStatus tag


# -----------------------------------------------------------------------------
# Main parser class
# -----------------------------------------------------------------------------

class OdpCaseParser:
    """Convert a Gherkin feature file into structured Excel rows."""

    def __init__(self, feature_dir: Path, feature_file: str):
        self.feature_dir = feature_dir
        self.feature_file = feature_file
        self.feature = FeatureParser(self.feature_dir, self.feature_file)

    # ------------------------------------------------------------------
    #  public
    # ------------------------------------------------------------------
    def parse(self):
        return self.feature.parse()

    # ------------------------------------------------------------------
    #  static helpers: step → text
    # ------------------------------------------------------------------
    @staticmethod
    def format_step(step: Step) -> str:
        """Turn a pytest‑bdd Step into multi‑line printable text."""
        base_text = f"{step.keyword} {step.name}".strip()

        table_lines: List[str] = []
        if step.datatable:
            for row in step.datatable.rows:
                table_lines.append("| " + " | ".join(cell.value for cell in row.cells) + " |")

        doc_lines = textwrap.indent(step.docstring, "  ") if step.docstring else ""

        blocks = [base_text] + table_lines + ([doc_lines] if doc_lines else [])
        return "\n".join(blocks)

    # ------------------------------------------------------------------
    #  Background 处理
    # ------------------------------------------------------------------
    @staticmethod
    def extract_background_data(bg: Background | None) -> Dict[str, str]:
        if not bg:
            return {}
        return {
            f"background_step_{idx}": OdpCaseParser.format_step(step)
            for idx, step in enumerate(bg.steps, start=1)
        }

    # ------------------------------------------------------------------
    #  Given/When/Then 分组
    # ------------------------------------------------------------------
    @staticmethod
    def collect_steps(
        steps: List[Step],
        bg_values: Set[str],
    ) -> Tuple[List[Tuple[int, str]], Dict[int, str]]:
        ordered_steps: List[Tuple[int, str]] = []
        expected: Dict[int, str] = {}
        seq = 0
        current_when: Optional[int] = None

        for step in steps:
            txt = OdpCaseParser.format_step(step)
            stype = step.type.lower()

            if stype == "given":
                if txt in bg_values:
                    continue  # 跳过来自 Background 的 Given
                seq += 1
                ordered_steps.append((seq, txt))

            elif stype == "when":
                seq += 1
                ordered_steps.append((seq, txt))
                current_when = seq

            elif stype == "then" and current_when is not None:
                expected[current_when] = (
                    expected.get(current_when, "") + ("\n" if current_when in expected else "") + txt
                )

        return ordered_steps, expected

    # ------------------------------------------------------------------
    #  Tag helpers
    # ------------------------------------------------------------------
    @staticmethod
    def parse_tags(tags: List[str]) -> Tuple[str, str, str]:
        """Return (fr_id, ccd_status, ncd_status) from a tag list."""
        fr_tags: List[str] = []
        ccd_status = ""
        ncd_status = ""
        for tag in tags or []:
            lowered = tag.lower()
            if lowered.startswith("odpcl."):
                fr_tags.append(f"[{tag}]")
            elif lowered.startswith("ccdbatchstatus="):
                ccd_status = tag.split("=", 1)[1]
            elif lowered.startswith("ncdbatchstatus="):
                ncd_status = tag.split("=", 1)[1]
        return "\n".join(fr_tags), ccd_status, ncd_status

    # ------------------------------------------------------------------
    #  Scenario → ParsedTestCase (normal & outline)
    # ------------------------------------------------------------------
    @staticmethod
    def _build_case(
        name: str,
        bg_data: Dict[str, str],
        bg_val_set: Set[str],
        steps: List[Step],
        tags: List[str],
    ) -> ParsedTestCase:
        ordered, expected = OdpCaseParser.collect_steps(steps, bg_val_set)
        fr_id, ccd, ncd = OdpCaseParser.parse_tags(tags)
        return ParsedTestCase(
            case_name=name,
            test_data=dict(bg_data),
            ordered_test_steps=ordered,
            expected_results=expected,
            fr_id=fr_id,
            ccd_batch_status=ccd,
            ncd_batch_status=ncd,
        )

    def _convert_scenario(self, sc: ScenarioTemplate, bg_data: Dict[str, str], bg_val_set: Set[str]) -> ParsedTestCase:
        return self._build_case(sc.name, bg_data, bg_val_set, sc.steps, sc.tags)

    def _convert_outline(self, tpl: ScenarioTemplate, bg_data: Dict[str, str], bg_val_set: Set[str]) -> List[ParsedTestCase]:
        cases: List[ParsedTestCase] = []
        for ex_block in tpl.examples:
            for ctx in ex_block.as_contexts():
                rendered = tpl.render(ctx)
                identifier = (
                    ctx.get("case identifier")
                    or ctx.get("case_identifier")
                    or ctx.get("case_name_identifier")
                )
                full_name = f"{rendered.name}-{identifier}" if identifier else rendered.name
                cases.append(
                    self._build_case(full_name, bg_data, bg_val_set, rendered.steps, tpl.tags)
                )
        return cases

    # ------------------------------------------------------------------
    #  Main orchestrator
    # ------------------------------------------------------------------
    def build_testcases_from_feature(self) -> List[ParsedTestCase]:
        feature = self.parse()
        bg_data = self.extract_background_data(feature.background)
        bg_val_set = set(bg_data.values())

        cases: List[ParsedTestCase] = []
        for _, sc_tpl in feature.scenarios.items():
            if sc_tpl.examples:
                cases.extend(self._convert_outline(sc_tpl, bg_data, bg_val_set))
            else:
                cases.append(self._convert_scenario(sc_tpl, bg_data, bg_val_set))
        return cases

    # ------------------------------------------------------------------
    #  Excel helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _fmt_bg_data(bg_data: Dict[str, str]) -> str:
        lines: List[str] = []
        for key in sorted(bg_data, key=lambda k: int(k.split("_")[-1])):
            val = bg_data[key]
            if val.startswith("Given "):
                val = val[len("Given "):]
            idx = key.split("_")[-1]
            lines.append(f"{idx}. {val}")
        return "\n".join(lines)

    @staticmethod
    def _auto_width(ws, headers: List[str], buffer: int = 1):
        for cell in ws[1]:
            if cell.value in headers:
                col = get_column_letter(cell.column)
                max_len = 0
                for c in ws[col]:
                    if c.value:
                        max_len = max(max_len, max(len(line) for line in str(c.value).splitlines()))
                ws.column_dimensions[col].width = max_len + buffer

    @staticmethod
    def write_testcases_to_excel(cases: List[ParsedTestCase], outfile: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "ParsedTestCases"

        headers = [
            "Test Case Name",
            "Test Data",
            "Test Steps",
            "Expected Results",
            "FR_ID",
            "CCDBatchStatus",
            "NCDBatchStatus",
        ]
        for idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=idx, value=h)

        row = 2
        for tc in cases:
            ws.cell(row=row, column=1, value=tc.case_name)
            ws.cell(row=row, column=2, value=OdpCaseParser._fmt_bg_data(tc.test_data))
            ws.cell(row=row, column=3, value="\n".join(f"{s}. {txt}" for s, txt in tc.ordered_test_steps))
            ws.cell(row=row, column=4, value="\n".join(f"{s}. {tc.expected_results[s]}" for s in sorted(tc.expected_results)))
            ws.cell(row=row, column=5, value=tc.fr_id)
            ws.cell(row=row, column=6, value=tc.ccd_batch_status)
            ws.cell(row=row, column=7, value=tc.ncd_batch_status)
            # wrap text columns
            for col in (2, 3, 4, 5):
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
            row += 1

        OdpCaseParser._auto_width(ws, ["Test Data", "Test Steps", "Expected Results", "FR_ID"])
        wb.save(outfile)


if __name__ == '__main__':
    parser = OdpCaseParser(Path(__file__).parent, "more_case.feature")
    build_result = parser.build_testcases_from_feature()
    parser.write_testcases_to_excel(build_result, "test_case11.xlsx")